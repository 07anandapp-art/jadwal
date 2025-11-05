import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import os

warna_bg_pic = {
    "Ananda": "#FF6F61",   # coral merah
    "Komar": "#6EC6FF",    # biru langit
    "Dante": "#81C784",    # hijau segar
    "Irgi": "#BA68C8",     # ungu terang
    "Aisyah": "#FFD54F",   # kuning cerah
    "Dilla": "#FFB74D",    # oranye lembut
    "Erit": "#A1887F",     # coklat abu
    "Syaiful": "#4DB6AC",  # turquoise
    "Arie": "#9575CD"      # lavender bold
}

kontak_pic = {
    "Ananda": "+62 821-7636-6808",
    "Komar": "+62 856-8787-796",
    "Dante": "+62 822-3333-2919",
    "Irgi": "+62 821-1597-2476",
    "Aisyah": "+62 851-2104-9424",
    "Dilla": "+62 851-2106-9312",
    "Erit": "+62 822-9921-3304",
    "Syaiful": "+62 851-5651-5205",
    "Arie": "+62 822-3199-3770"
}

st.set_page_config(page_title="Kalender Jadwal Uangel", layout="wide")

if "reset_detail" not in st.session_state:
    st.session_state.reset_detail = ""

# --- Load data jadwal ---
if os.path.exists("jadwal.csv"):
    df = pd.read_csv("jadwal.csv")
    df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce")
else:
    df = pd.DataFrame(columns=["tanggal", "jenis", "pic", "aktivitas", "jam", "detail"])

# --- Sidebar Input Jadwal ---
st.sidebar.header("Input Jadwal")
jenis = st.sidebar.selectbox("Jenis", ["Libur Nasional", "Standby"])
tanggal = st.sidebar.date_input("Tanggal", value=datetime.today())
detail_aktivitas = st.sidebar.text_area("Detail aktivitas (opsional)", key="detail_input")

# Dropdown PIC tetap
daftar_pic = ["Ananda", "Komar", "Dante", "Irgi", "Aisyah", "Dilla", "Erit", "Syaiful", "Arie"]
pic = st.sidebar.selectbox("PIC", daftar_pic) if jenis == "Standby" else "-"

# Input aktivitas
aktivitas_opsi = ["Standby Naru", "Standby Lebaran", "Activity Malam DRA"]
aktivitas = st.sidebar.selectbox("Nama aktivitas", aktivitas_opsi)

# Dropdown jam standby (opsional)
jam_opsi = ["", "08:00-16:00", "16:00-00:00", "00:00-08:00"]
jam_standby = st.sidebar.selectbox("Jam Standby (opsional)", jam_opsi) if jenis == "Standby" else "-"

# Tombol tambah
if st.sidebar.button("Tambah"):
    if aktivitas.strip() == "":
        st.sidebar.warning("Nama aktivitas tidak boleh kosong.")
    else:
        if jenis == "Standby" and jam_standby.strip() == "":
            jam_standby = "-"
        
        new_entry = {
            "tanggal": tanggal.strftime("%Y-%m-%d"),
            "jenis": jenis,
            "pic": pic,
            "aktivitas": aktivitas,
            "jam": jam_standby,
            "detail": detail_aktivitas.strip(),
            "sumber": "manual"

        }
        df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
        df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce")
        df.to_csv("jadwal.csv", index=False)
        st.sidebar.success(f"Jadwal {jenis} berhasil ditambahkan untuk {tanggal.strftime('%d %b %Y')}")

        st.rerun()  # ✅ refresh halaman, input kosong otomatis

        # Kosongkan input dan refresh halaman
        st.session_state["detail_input"] = ""
        st.experimental_rerun()

        # Tambahkan ke DataFrame
        df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
        df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce")
        df.to_csv("jadwal.csv", index=False)
        st.sidebar.success(f"Jadwal {jenis} berhasil ditambahkan untuk {tanggal.strftime('%d %b %Y')}")
        
        # Reset kolom detail setelah tambah
        st.session_state["detail_input"] = ""

# --- Pilih Bulan ---
st.title("Kalender Jadwal Uangel")
bulan_list = pd.date_range("2025-01-01", "2026-12-01", freq="MS")
bulan_str_list = bulan_list.strftime("%B %Y")
bulan_sekarang = datetime.today().strftime("%B %Y")
index_default = bulan_str_list.tolist().index(bulan_sekarang) if bulan_sekarang in bulan_str_list.tolist() else 0

bulan_terpilih = st.selectbox("Pilih Bulan", bulan_str_list, index=index_default)
bulan_dt = pd.to_datetime(bulan_terpilih, format="%B %Y")
start_date = bulan_dt
end_date = (bulan_dt + pd.DateOffset(months=1)) - timedelta(days=1)
dates = pd.date_range(start_date, end_date)

# --- Filter jadwal untuk bulan terpilih ---
jadwal_bulan = df[(df["tanggal"] >= start_date) & (df["tanggal"] <= end_date)]

# --- Tampilkan Jadwal dalam Grid HTML ---
# Nama hari dalam bahasa Indonesia
nama_hari = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]

# Header hari
html = "<div style='display: grid; grid-template-columns: repeat(7, 1fr); gap: 5px; font-weight: bold;'>"
for hari in nama_hari:
    html += f"<div style='text-align:center;'>{hari}</div>"
html += "</div>"

# Grid tanggal
html += "<div style='display: grid; grid-template-columns: repeat(7, 1fr); gap: 5px;'>"

# Hitung offset awal (misalnya jika bulan dimulai hari Rabu)
offset = pd.to_datetime(dates[0]).weekday()
for _ in range(offset):
    html += "<div></div>"

for date in dates:
    jadwal = jadwal_bulan[jadwal_bulan["tanggal"] == date]
    warna = "#f0f0f0"
    isi = f"<b>{date.strftime('%d %b')}</b>"

    # Deteksi hari Sabtu/Minggu
    if date.weekday() in [5, 6]:  # 5=Sabtu, 6=Minggu
        warna = "#FFEEEE"

    if not jadwal.empty:
        if "Libur Nasional" in jadwal["jenis"].values:
            warna = "#FFCCCC"
        else:
            warna = "#CCFFCC"  # hijau untuk manual
        # Gabungkan berdasarkan aktivitas
        aktivitas_group = jadwal.groupby("aktivitas")

        for aktivitas, grup in aktivitas_group:
            isi += f"<br><b>{aktivitas}</b>"

            # Urutkan berdasarkan jam kerja
            def jam_sort_key(jam):
                if jam == "-" or not isinstance(jam, str):
                    return 999  # taruh paling bawah
                try:
                    return int(jam.split(":")[0])  # ambil jam awal sebagai kunci
                except:
                    return 999
            grup_sorted = grup.sort_values(by="jam", key=lambda col: col.map(jam_sort_key))

            for _, row in grup_sorted.iterrows():
                if row.get("sumber") == "acak":
                    warna_bg = warna_bg_pic.get(row["pic"], "#EEEEEE")
                    isi += f"<div style='margin-top:2px; background:{warna_bg}; padding:2px 4px; border-radius:3px;'>{row['pic']} {row['jam']}</div>"
                else:
                    isi += f"<div style='margin-top:2px;'>{row['pic']} {row['jam']}</div>"

                

    html += f"""
<div style='
    background:{warna};
    padding:8px;
    border:1px solid #ccc;
    border-radius:5px;
    min-height:100px;
    max-height:none;
    overflow: visible;
    word-wrap:break-word;
    font-size: clamp(9px, 1vw, 11px);
    line-height: 1.3;
'>
{isi}
</div>
"""

html += "</div>"
st.markdown(html, unsafe_allow_html=True)

# Tabel + Hapus Jadwal
st.subheader("Daftar Jadwal Bulan Ini")
jadwal_bulan_reset = jadwal_bulan.reset_index(drop=True)

if not jadwal_bulan_reset.empty:
    kolom_urut = ["tanggal", "jenis", "pic", "aktivitas", "jam", "detail"]
    st.dataframe(jadwal_bulan_reset[kolom_urut])
    hapus_index = st.number_input("Masukkan nomor baris yang ingin dihapus", min_value=0, max_value=len(jadwal_bulan_reset)-1, step=1)
    if st.button("Hapus Jadwal"):
        index_asli = jadwal_bulan_reset.loc[hapus_index].name
        df.drop(index_asli, inplace=True)
        df.to_csv("jadwal.csv", index=False)
        st.success("Jadwal berhasil dihapus. Silakan refresh halaman.")
else:
    st.info("Belum ada jadwal di bulan ini.")

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

if not jadwal_bulan_reset.empty:
    if st.button("Export ke Excel"):
        # Buat file Excel
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Jadwal Bulan Ini"

        # Header
        header = ["Tanggal", "Jenis", "PIC", "Aktivitas", "Jam", "Detail"]
        ws.append(header)

        # Styling header
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, col_name in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Isi data
        for i, row in jadwal_bulan_reset.iterrows():
            ws.append([
                row["tanggal"].strftime("%Y-%m-%d") if pd.notnull(row["tanggal"]) else "",
                row["jenis"],
                row["pic"],
                row["aktivitas"],
                row["jam"],
                row["detail"]
            ])

        # Auto width
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Simpan file
        wb.save("jadwal_bulan_ini.xlsx")
        st.success("Jadwal berhasil diexport ke Excel sebagai 'jadwal_bulan_ini.xlsx'")
        
        
from datetime import datetime, timedelta
import pandas as pd
import random
import streamlit as st

# Load jadwal.csv jika ada
try:
    df = pd.read_csv("jadwal.csv")
    df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce")
except FileNotFoundError:
    df = pd.DataFrame(columns=["tanggal", "jenis", "pic", "aktivitas", "jam", "detail", "sumber"])

with st.expander("📅 Buat Jadwal Standby Uangel"):
    tanggal_mulai = st.date_input("Tanggal mulai", value=datetime.today())
    tanggal_selesai = st.date_input("Tanggal selesai", value=datetime.today() + timedelta(days=30))
    daftar_pic = st.text_input("Daftar PIC (pisahkan dengan koma)", value="Ananda,Komar,Dante,Irgi,Aisyah,Dilla,Erit,Syaiful,Arie")
    jam_opsi = ["00:00 - 08:00", "08:00 - 16:00", "16:00 - 00:00"]
    jam_standby_list = st.multiselect("Pilih jam standby", jam_opsi, default=jam_opsi)
    aktivitas_default = st.text_input("Nama aktivitas", value="Standby Tim")
    batas_per_pic_input = st.text_area("Batas maksimal standby per PIC (opsional)\nContoh: Ananda=5,Komar=7", value="")

    # Parsing batas maksimal
    batas_per_pic = {}
    if batas_per_pic_input.strip():
        for item in batas_per_pic_input.split(","):
            if "=" in item:
                nama, batas = item.split("=")
                batas_per_pic[nama.strip()] = int(batas.strip())

    if st.button("Generate Jadwal Acak"):
        pic_list = [p.strip() for p in daftar_pic.split(",") if p.strip()]
        tanggal_range = pd.date_range(tanggal_mulai, tanggal_selesai).tolist()
        jadwal_baru = []
        counter_pic = {pic: 0 for pic in pic_list}

        def is_continuous(prev_jadwal, tanggal, jam, pic):
            jam = jam.strip()
            tanggal_str = tanggal.strftime("%Y-%m-%d")
            tanggal_prev = (tanggal - timedelta(days=1)).strftime("%Y-%m-%d")

            for entry in prev_jadwal:
                if entry["pic"] != pic:
                    continue
                if jam == "00:00 - 08:00" and entry["tanggal"] == tanggal_prev and entry["jam"] == "16:00 - 00:00":
                    return True
                if jam == "08:00 - 16:00" and entry["tanggal"] == tanggal_str and entry["jam"] == "00:00 - 08:00":
                    return True
                if jam == "16:00 - 00:00" and entry["tanggal"] == tanggal_str and entry["jam"] == "08:00 - 16:00":
                    return True
            return False

        for tanggal in tanggal_range:
            for jam in jam_standby_list:
                random.shuffle(pic_list)
        
                # Cari kandidat PIC yang lolos aturan dan belum melebihi batas
                kandidat_pic = []
                for pic in pic_list:
                    if pic in batas_per_pic and counter_pic[pic] >= batas_per_pic[pic]:
                        continue
                    if is_continuous(jadwal_baru, tanggal, jam, pic):
                        continue
                    kandidat_pic.append(pic)
        
                # Urutkan berdasarkan jumlah shift terkecil
                kandidat_pic.sort(key=lambda p: counter_pic[p])
        
                # Acak di antara kandidat dengan jumlah shift terkecil
                if kandidat_pic:
                    min_shift = counter_pic[kandidat_pic[0]]
                    kandidat_terpilih = [p for p in kandidat_pic if counter_pic[p] == min_shift]
                    pic_final = random.choice(kandidat_terpilih)
        
                    jadwal_baru.append({
                        "tanggal": tanggal.strftime("%Y-%m-%d"),
                        "jenis": "Standby",
                        "pic": pic_final,
                        "aktivitas": aktivitas_default,
                        "jam": jam,
                        "detail": "",
                        "sumber": "acak"  # 🔶 penanda untuk pewarnaan kalender
                    })
                    counter_pic[pic_final] += 1
                    

        # Simpan ke CSV
        df = pd.concat([df, pd.DataFrame(jadwal_baru)], ignore_index=True)
        df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce")
        df.to_csv("jadwal.csv", index=False)
        st.success(f"{len(jadwal_baru)} jadwal standby berhasil dibuat dan disimpan!")

def generate_export_html(df, bulan_dt):
    start_date = bulan_dt
    end_date = (bulan_dt + pd.DateOffset(months=1)) - timedelta(days=1)
    dates = pd.date_range(start_date, end_date)
    jadwal_bulan = df[(df["tanggal"] >= start_date) & (df["tanggal"] <= end_date)]
    html = f"""
    <div style='background:white; padding:20px; font-family:sans-serif; color:black;'>
    <h2>Kalender Jadwal Uangel - {bulan_dt.strftime('%B %Y')}</h2>
    """
    # Kalender Grid
    nama_hari = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    html += "<div style='display: grid; grid-template-columns: repeat(7, 1fr); gap: 5px;'>"

    offset = pd.to_datetime(dates[0]).weekday()
    for _ in range(offset):
        html += "<div></div>"

    for date in dates:
        jadwal = jadwal_bulan[jadwal_bulan["tanggal"] == date]
        warna = "#f0f0f0"
        isi = f"<b>{date.strftime('%d %b')}</b>"

        if date.weekday() in [5, 6]:
            warna = "#FFEEEE"
        if not jadwal.empty:
            if "Libur Nasional" in jadwal["jenis"].values:
                warna = "#FFCCCC"
            else:
                warna = "#CCFFCC"

            aktivitas_group = jadwal.groupby("aktivitas")
            for aktivitas, grup in aktivitas_group:
                isi += f"<br><b>{aktivitas}</b>"
                grup_sorted = grup.sort_values(by="jam")
                for _, row in grup_sorted.iterrows():
                    if row.get("sumber") == "acak":
                        warna_bg = warna_bg_pic.get(row["pic"], "#EEEEEE")
                        isi += f"<div style='margin-top:2px; background:{warna_bg}; padding:2px 4px; border-radius:3px;'>{row['pic']} {row['jam']}</div>"
                    else:
                        isi += f"<div style='margin-top:2px;'>{row['pic']} {row['jam']}</div>"

        html += f"<div style='background:{warna}; padding:8px; border:1px solid #ccc; border-radius:5px; min-height:100px;'>{isi}</div>"

    # Summary Shift Acak
    df_acak = jadwal_bulan[jadwal_bulan["sumber"] == "acak"]
    summary = df_acak["pic"].value_counts().reset_index()
    summary.columns = ["PIC", "Jumlah Shift"]
    html += "<h3>📊 Summary Jadwal Standby Uangel</h3><ul>"
    for _, row in summary.iterrows():
        html += f"<li style='white-space:nowrap;'>{row['PIC']}: {row['Jumlah Shift']} shift</li>"
    html += "</ul>"

    # Kontak PIC
    html += "<h3>📞 Kontak PIC</h3><ul>"
    for nama, nomor in kontak_pic.items():
        html += f"<li style='white-space:nowrap;'>{nama}: {nomor}</li>"
    html += "</ul>"
    html += "<div style='height:50px;'></div>"  # spacer bawah
    html += "</div>"

    return html


from html2image import Html2Image
def export_jadwal_to_png(html_content, filename="jadwal_tim.png"):
    hti = Html2Image()
    hti.output_path = "."  # simpan di folder saat ini
    hti.screenshot(html_str=html_content, save_as=filename, size=(1200, 1600))
    
# --- Tombol Export PNG ---
st.markdown("---")
st.markdown("### 🖼️ Export Jadwal ke Gambar PNG")

if st.button("Export Kalender + Summary ke PNG"):
    html_export = generate_export_html(df, bulan_dt)
    export_jadwal_to_png(html_export)
    st.success("Gambar berhasil dibuat sebagai 'jadwal_tim.png'")